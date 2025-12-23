import csv
from openpyxl import load_workbook

def convert_xlsx_to_csv(xlsx_path: str, csv_path: str, sheet_name: str | None = None):
    """
    Streams Excel (.xlsx) into CSV WITHOUT loading entire workbook into memory.
    This is critical for large files (100MB+).
    """

    wb = load_workbook(
        filename=xlsx_path,
        read_only=True,      # prevents memory explosion
        data_only=True       # resolves formulas -> values
    )

    # pick sheet
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found.")
        sheet = wb[sheet_name]
    else:
        sheet = wb[wb.sheetnames[0]]  # first sheet

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        for row in sheet.iter_rows(values_only=True):
            # convert None → ""
            cleaned_row = ["" if v is None else v for v in row]
            writer.writerow(cleaned_row)

    wb.close()
    return csv_path
